"""
AgTools Data Export Service
Exports data to CSV and Excel formats
"""

import csv
import io
from datetime import datetime
from typing import Any, Dict, List, Optional
from dataclasses import dataclass
from enum import Enum


class ExportFormat(str, Enum):
    CSV = "csv"
    EXCEL = "excel"


@dataclass
class ExportConfig:
    """Configuration for data export"""
    include_headers: bool = True
    date_format: str = "%Y-%m-%d"
    datetime_format: str = "%Y-%m-%d %H:%M:%S"
    decimal_places: int = 2
    include_metadata: bool = True
    sheet_name: str = "Data"


class DataExportService:
    """Service for exporting data to various formats"""

    def __init__(self):
        self.default_config = ExportConfig()

    def _format_value(self, value: Any, config: ExportConfig) -> Any:
        """Format a value for export"""
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime(config.datetime_format)
        if isinstance(value, float):
            return round(value, config.decimal_places)
        if isinstance(value, bool):
            return "Yes" if value else "No"
        if isinstance(value, (list, dict)):
            return str(value)
        return value

    def _prepare_rows(
        self,
        data: List[Dict[str, Any]],
        columns: Optional[List[str]] = None,
        config: Optional[ExportConfig] = None
    ) -> tuple:
        """Prepare data rows for export"""
        config = config or self.default_config

        if not data:
            return [], []

        # Determine columns
        if columns:
            headers = columns
        else:
            # Get all unique keys from data
            headers = []
            for row in data:
                for key in row.keys():
                    if key not in headers:
                        headers.append(key)

        # Format rows
        rows = []
        for row in data:
            formatted_row = []
            for col in headers:
                value = row.get(col, "")
                formatted_row.append(self._format_value(value, config))
            rows.append(formatted_row)

        return headers, rows

    def export_to_csv(
        self,
        data: List[Dict[str, Any]],
        columns: Optional[List[str]] = None,
        config: Optional[ExportConfig] = None
    ) -> bytes:
        """Export data to CSV format"""
        config = config or self.default_config
        headers, rows = self._prepare_rows(data, columns, config)

        output = io.StringIO()
        writer = csv.writer(output)

        if config.include_headers and headers:
            # Make headers more readable
            readable_headers = [h.replace("_", " ").title() for h in headers]
            writer.writerow(readable_headers)

        for row in rows:
            writer.writerow(row)

        return output.getvalue().encode('utf-8-sig')  # BOM for Excel compatibility

    def export_to_excel(
        self,
        data: List[Dict[str, Any]],
        columns: Optional[List[str]] = None,
        config: Optional[ExportConfig] = None
    ) -> bytes:
        """Export data to Excel format"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

        config = config or self.default_config
        headers, rows = self._prepare_rows(data, columns, config)

        wb = Workbook()
        ws = wb.active
        ws.title = config.sheet_name

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write headers
        if config.include_headers and headers:
            readable_headers = [h.replace("_", " ").title() for h in headers]
            for col_num, header in enumerate(readable_headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

        # Write data
        start_row = 2 if config.include_headers else 1
        for row_num, row in enumerate(rows, start_row):
            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                # Right-align numbers
                if isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal="right")

        # Auto-adjust column widths
        for col_num, header in enumerate(headers, 1):
            max_length = len(str(header))
            for row in rows:
                if col_num <= len(row):
                    cell_value = str(row[col_num - 1])
                    max_length = max(max_length, len(cell_value))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[get_column_letter(col_num)].width = adjusted_width

        # Add metadata if requested
        if config.include_metadata:
            # Add export timestamp in a separate row at the bottom
            last_row = len(rows) + start_row + 1
            ws.cell(row=last_row, column=1, value=f"Exported: {datetime.now().strftime(config.datetime_format)}")

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    # =========================================================================
    # Pre-configured Export Methods
    # =========================================================================

    def export_fields(
        self,
        fields: List[Dict],
        format: ExportFormat = ExportFormat.CSV
    ) -> bytes:
        """Export field data"""
        columns = [
            "name", "acres", "crop", "soil_type", "irrigation_type",
            "planting_date", "notes", "status"
        ]
        config = ExportConfig(sheet_name="Fields")

        if format == ExportFormat.EXCEL:
            return self.export_to_excel(fields, columns, config)
        return self.export_to_csv(fields, columns, config)

    def export_operations(
        self,
        operations: List[Dict],
        format: ExportFormat = ExportFormat.CSV
    ) -> bytes:
        """Export field operations data"""
        columns = [
            "field_name", "operation_type", "date", "operator",
            "equipment", "notes", "cost", "status"
        ]
        config = ExportConfig(sheet_name="Operations")

        if format == ExportFormat.EXCEL:
            return self.export_to_excel(operations, columns, config)
        return self.export_to_csv(operations, columns, config)

    def export_equipment(
        self,
        equipment: List[Dict],
        format: ExportFormat = ExportFormat.CSV
    ) -> bytes:
        """Export equipment inventory"""
        columns = [
            "name", "type", "make", "model", "year", "serial_number",
            "current_hours", "status", "location", "purchase_price"
        ]
        config = ExportConfig(sheet_name="Equipment")

        if format == ExportFormat.EXCEL:
            return self.export_to_excel(equipment, columns, config)
        return self.export_to_csv(equipment, columns, config)

    def export_maintenance(
        self,
        maintenance: List[Dict],
        format: ExportFormat = ExportFormat.CSV
    ) -> bytes:
        """Export maintenance records"""
        columns = [
            "equipment_name", "maintenance_type", "date", "hours_at_service",
            "cost", "parts_used", "technician", "notes", "next_due_date"
        ]
        config = ExportConfig(sheet_name="Maintenance")

        if format == ExportFormat.EXCEL:
            return self.export_to_excel(maintenance, columns, config)
        return self.export_to_csv(maintenance, columns, config)

    def export_inventory(
        self,
        inventory: List[Dict],
        format: ExportFormat = ExportFormat.CSV
    ) -> bytes:
        """Export inventory items"""
        columns = [
            "name", "category", "quantity", "unit", "unit_cost",
            "total_value", "reorder_point", "location", "expiration_date"
        ]
        config = ExportConfig(sheet_name="Inventory")

        if format == ExportFormat.EXCEL:
            return self.export_to_excel(inventory, columns, config)
        return self.export_to_csv(inventory, columns, config)

    def export_tasks(
        self,
        tasks: List[Dict],
        format: ExportFormat = ExportFormat.CSV
    ) -> bytes:
        """Export task list"""
        columns = [
            "title", "description", "assigned_to", "field_name",
            "priority", "status", "due_date", "completed_date", "notes"
        ]
        config = ExportConfig(sheet_name="Tasks")

        if format == ExportFormat.EXCEL:
            return self.export_to_excel(tasks, columns, config)
        return self.export_to_csv(tasks, columns, config)

    def export_costs(
        self,
        costs: List[Dict],
        format: ExportFormat = ExportFormat.CSV
    ) -> bytes:
        """Export cost tracking data"""
        columns = [
            "date", "category", "description", "field_name",
            "quantity", "unit", "unit_cost", "total_cost", "vendor"
        ]
        config = ExportConfig(sheet_name="Costs")

        if format == ExportFormat.EXCEL:
            return self.export_to_excel(costs, columns, config)
        return self.export_to_csv(costs, columns, config)

    def export_cost_per_acre(
        self,
        data: List[Dict],
        format: ExportFormat = ExportFormat.CSV
    ) -> bytes:
        """Export cost per acre report"""
        columns = [
            "field_name", "crop", "acres", "seed_cost", "fertilizer_cost",
            "chemical_cost", "fuel_cost", "labor_cost", "other_cost",
            "total_cost", "cost_per_acre"
        ]
        config = ExportConfig(sheet_name="Cost Per Acre")

        if format == ExportFormat.EXCEL:
            return self.export_to_excel(data, columns, config)
        return self.export_to_csv(data, columns, config)

    def export_profitability(
        self,
        data: List[Dict],
        format: ExportFormat = ExportFormat.CSV
    ) -> bytes:
        """Export profitability report"""
        columns = [
            "field_name", "crop", "acres", "yield_bu", "price_per_bu",
            "gross_revenue", "total_costs", "net_profit", "profit_per_acre",
            "roi_percent"
        ]
        config = ExportConfig(sheet_name="Profitability")

        if format == ExportFormat.EXCEL:
            return self.export_to_excel(data, columns, config)
        return self.export_to_csv(data, columns, config)

    def export_spray_records(
        self,
        records: List[Dict],
        format: ExportFormat = ExportFormat.CSV
    ) -> bytes:
        """Export spray application records"""
        columns = [
            "date", "field_name", "crop", "target_pest", "product_name",
            "rate", "acres_treated", "total_product", "cost",
            "applicator", "weather_conditions", "wind_speed", "temperature"
        ]
        config = ExportConfig(sheet_name="Spray Records")

        if format == ExportFormat.EXCEL:
            return self.export_to_excel(records, columns, config)
        return self.export_to_csv(records, columns, config)

    def export_scouting_records(
        self,
        records: List[Dict],
        format: ExportFormat = ExportFormat.CSV
    ) -> bytes:
        """Export scouting observation records"""
        columns = [
            "date", "field_name", "crop", "growth_stage", "scout_name",
            "observation_type", "finding", "severity", "location",
            "recommended_action", "photo_count"
        ]
        config = ExportConfig(sheet_name="Scouting")

        if format == ExportFormat.EXCEL:
            return self.export_to_excel(records, columns, config)
        return self.export_to_csv(records, columns, config)

    def export_multi_sheet(
        self,
        sheets: Dict[str, List[Dict]],
        columns_map: Optional[Dict[str, List[str]]] = None
    ) -> bytes:
        """Export multiple data sets to a single Excel workbook with multiple sheets"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("openpyxl is required for Excel export")

        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for sheet_name, data in sheets.items():
            ws = wb.create_sheet(title=sheet_name[:31])  # Excel limits sheet names to 31 chars

            columns = columns_map.get(sheet_name) if columns_map else None
            headers, rows = self._prepare_rows(data, columns, self.default_config)

            if not headers:
                continue

            # Write headers
            readable_headers = [h.replace("_", " ").title() for h in headers]
            for col_num, header in enumerate(readable_headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # Write data
            for row_num, row in enumerate(rows, 2):
                for col_num, value in enumerate(row, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.border = thin_border

            # Auto-adjust column widths
            for col_num, header in enumerate(headers, 1):
                max_length = len(str(header))
                for row in rows:
                    if col_num <= len(row):
                        max_length = max(max_length, len(str(row[col_num - 1])))
                ws.column_dimensions[get_column_letter(col_num)].width = min(max_length + 2, 50)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()


# Singleton instance
_export_service = None


def get_data_export_service() -> DataExportService:
    """Get or create the data export service singleton"""
    global _export_service
    if _export_service is None:
        _export_service = DataExportService()
    return _export_service
